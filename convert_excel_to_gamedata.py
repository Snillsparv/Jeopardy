#!/usr/bin/env python3
"""
Konverterar frågor_jeopardy_2026.xlsx till gamedata.js
- Ingen svar behövs
- Daily Doubles: 1 i Round 1, 2 i Round 2, 3 i Round 3
- [text] → bildväg images/questions/text.png
- Intro: → ljudväg sounds/intros/låtnamn.mp3
"""

import openpyxl
import random
import json
import re

def slugify(text):
    """Konverterar text till filnamn-vänlig sträng"""
    text = text.lower()
    text = re.sub(r'[åä]', 'a', text)
    text = re.sub(r'ö', 'o', text)
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return text.strip('_')

def process_question(question_text):
    """
    Bearbetar frågetext för bilder och ljud
    [text] → <img src="images/questions/text.png">
    Intro: låtnamn → <audio src="sounds/intros/latnamn.mp3">
    """
    if not question_text:
        return "", None, None

    media_files = []

    # Kolla efter bildfrågor [text]
    bracket_match = re.search(r'\[(.*?)\]', question_text)
    if bracket_match:
        image_name = bracket_match.group(1)
        image_slug = slugify(image_name)
        image_path = f"images/questions/{image_slug}.png"
        media_files.append(('image', image_path, image_name))
        # Ersätt [text] med bildtagg
        question_text = re.sub(r'\[.*?\]', f'<img src="{image_path}" alt="{image_name}" style="max-width: 600px; max-height: 400px; margin: 20px 0;">', question_text)

    # Kolla efter intro-frågor (med eller utan kolon)
    intro_match = re.match(r'Intro:?\s*(.*)', question_text, re.IGNORECASE)
    if intro_match:
        song_name = intro_match.group(1).strip()
        song_slug = slugify(song_name)
        audio_path = f"sounds/intros/{song_slug}.mp3"
        media_files.append(('audio', audio_path, song_name))
        # Ersätt med ljudtagg
        question_text = f'<div style="text-align: center;"><p style="font-size: 2rem; margin-bottom: 20px;">Lyssna på introt:</p><audio controls autoplay src="{audio_path}"></audio></div>'

    return question_text, media_files

def parse_round(sheet):
    """Parsar ett ark och returnerar kategorier och frågor"""
    categories = []
    questions = []
    media_files = []

    current_category = None
    category_questions = []

    for row in sheet.iter_rows(values_only=True):
        # Kolla om det är en ny kategori (kolumn 1 har text)
        if row[0]:
            # Spara föregående kategori om den finns
            if current_category and category_questions:
                categories.append(current_category)
                questions.append(category_questions)

            # Starta ny kategori
            current_category = row[0]
            category_questions = []

        # Kolla om det finns en fråga (kolumn 3)
        if row[2]:
            processed_question, media = process_question(row[2])
            if media:
                media_files.extend(media)
            category_questions.append(processed_question)

    # Lägg till sista kategorin
    if current_category and category_questions:
        categories.append(current_category)
        questions.append(category_questions)

    return categories, questions, media_files

def generate_daily_doubles(num_categories, num_questions_per_category, count):
    """Slumpar Daily Double-positioner"""
    positions = []
    for _ in range(count):
        col = random.randint(0, num_categories - 1)
        row = random.randint(0, num_questions_per_category - 1)
        position = f"{col}-{row}"
        # Se till att vi inte får duplicates
        while position in positions:
            col = random.randint(0, num_categories - 1)
            row = random.randint(0, num_questions_per_category - 1)
            position = f"{col}-{row}"
        positions.append(position)
    return positions

def create_gamedata(rounds_data, media_files):
    """Skapar gamedata.js innehåll"""

    output = "// Jeopardy Game Data\n"
    output += "// Genererad från frågor_jeopardy_2026.xlsx\n\n"
    output += "const gameData = {\n"

    # Bearbeta varje omgång
    round_names = ['round1', 'round2', 'round3']
    daily_double_counts = [1, 2, 3]

    for i, (round_name, dd_count) in enumerate(zip(round_names, daily_double_counts)):
        categories, questions_by_category, round_media = rounds_data[i]

        output += f"    {round_name}: {{\n"

        # Kategorier
        output += f"        categories: {json.dumps(categories, ensure_ascii=False)},\n"

        # Frågor
        output += "        questions: [\n"

        # Skapa frågor med rätt värden (10, 20, 30, 40, 50 för Round 1, etc.)
        base_value = (i + 1) * 100
        values = [base_value, base_value * 2, base_value * 3, base_value * 4, base_value * 5]

        for cat_idx, cat_questions in enumerate(questions_by_category):
            output += "            [\n"
            for q_idx, question in enumerate(cat_questions):
                value = values[q_idx] if q_idx < len(values) else values[-1]
                # Escape quotes i frågetexten
                question_escaped = question.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n')
                output += f'                {{value: {value}, question: "{question_escaped}", answer: ""}},\n'
            output += "            ],\n"

        output += "        ],\n"

        # Daily Doubles
        num_cats = len(categories)
        num_qs = len(questions_by_category[0]) if questions_by_category else 5
        daily_doubles = generate_daily_doubles(num_cats, num_qs, dd_count)
        output += f"        dailyDoubles: {json.dumps(daily_doubles)}\n"

        output += "    },\n\n"

    # Final Jeopardy (placeholder)
    output += "    final: {\n"
    output += "        category: 'Final Jeopardy',\n"
    output += "        question: 'Final Jeopardy-fråga här',\n"
    output += "        answer: ''\n"
    output += "    }\n"
    output += "};\n"

    return output

def main():
    # Läs Excel-filen
    wb = openpyxl.load_workbook('/home/user/Jeopardy/frågor_jeopardy_2026.xlsx')

    rounds_data = []
    all_media_files = []

    # Bearbeta alla omgångar
    for sheet_name in ['Omgång 1', 'Omgång 2', 'Omgång 3']:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            categories, questions, media_files = parse_round(sheet)
            rounds_data.append((categories, questions, media_files))
            all_media_files.extend(media_files)

    # Generera gamedata.js
    gamedata_content = create_gamedata(rounds_data, all_media_files)

    # Skriv till fil
    with open('/home/user/Jeopardy/gamedata.js', 'w', encoding='utf-8') as f:
        f.write(gamedata_content)

    print("✅ gamedata.js skapad!")
    print(f"\n📊 Statistik:")
    print(f"   - {len(rounds_data)} omgångar")
    for i, (cats, qs, _) in enumerate(rounds_data, 1):
        total_questions = sum(len(q) for q in qs)
        print(f"   - Omgång {i}: {len(cats)} kategorier, {total_questions} frågor")

    # Skriv lista över mediafilar som behövs
    if all_media_files:
        print(f"\n📁 Mediafilar som behövs ({len(all_media_files)} st):")

        images = [f for f in all_media_files if f[0] == 'image']
        audios = [f for f in all_media_files if f[0] == 'audio']

        if images:
            print(f"\n🖼️  BILDER (lägg i images/questions/):")
            for _, path, name in images:
                filename = path.split('/')[-1]
                print(f"   - {filename}  (för: {name})")

        if audios:
            print(f"\n🎵 LJUDFILER (lägg i sounds/intros/):")
            for _, path, name in audios:
                filename = path.split('/')[-1]
                print(f"   - {filename}  (för: {name})")

if __name__ == '__main__':
    main()
